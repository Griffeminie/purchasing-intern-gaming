"""
clean_monitoring_excel.py

Operates directly on an uploaded monitoring.xlsx workbook (not the JSON cache).

Modes:
  inspect <path>            -> prints JSON stats about merged cells / phantom rows
  clean <path> <out_path>   -> unmerges all merged cells (filling values into every
                                cell in the range), trims phantom empty rows past the
                                real data, saves to out_path, prints JSON summary

Usage:
    python clean_monitoring_excel.py inspect data/uploads/pending_123.xlsx
    python clean_monitoring_excel.py clean data/uploads/pending_123.xlsx data/monitoring.xlsx
"""

import sys
import json
import argparse
import openpyxl


def scan_real_last_row(ws, scan_limit=500, log=None, sheet_name=""):
    """Find the last row that actually has content, scanning up to scan_limit rows.
    Logs periodic progress if a log() function is given.

    IMPORTANT: never scan beyond ws.max_row as it currently stands. Forcing
    iter_rows(max_row=N) where N > ws.max_row makes openpyxl materialize new
    empty row/cell objects up to N - and on some sheets (depending on styles,
    data validation, etc. baked into the template) that expansion has been
    observed to be catastrophically slow (minutes, not milliseconds) even
    though the sheet has almost no real data. Capping to the smaller of
    scan_limit and the sheet's own current max_row avoids ever triggering
    that expansion unnecessarily.
    """
    effective_limit = min(scan_limit, ws.max_row) if ws.max_row else 0
    if effective_limit == 0:
        return 0

    real_last = 0
    rows_with_content = 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=effective_limit, values_only=True), start=1):
        has_content = any(c is not None and str(c).strip() != "" for c in row)
        if has_content:
            real_last = i
            rows_with_content += 1
        if log and i % 500 == 0:
            log(f"sheet={sheet_name} scanning... row {i}/{effective_limit} (contentSoFar={rows_with_content}, lastContentRow={real_last})")
    return real_last


def inspect_workbook(path):
    import time as _time

    def log(msg):
        print(f"[progress] {msg}", file=sys.stderr, flush=True)

    t_start = _time.time()
    log(f"opening workbook: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    log(f"loaded in {_time.time() - t_start:.1f}s, found {len(wb.sheetnames)} sheets: {wb.sheetnames}")

    sheets = []
    total_merged = 0
    any_phantom = False

    # IMPORTANT: capture ws.max_row for every sheet FIRST, before calling
    # scan_real_last_row() on any sheet. openpyxl's iter_rows(max_row=N)
    # allocates cell objects up to N as a side effect, which inflates
    # ws.max_row afterwards if it hadn't been read yet. Reading it upfront
    # avoids that trap.
    reported_max_rows = {name: wb[name].max_row for name in wb.sheetnames}

    for name in wb.sheetnames:
        t_sheet = _time.time()
        ws = wb[name]
        log(f"sheet={name} scanning for merged cells...")
        merged_count = len(ws.merged_cells.ranges)
        total_merged += merged_count
        log(f"sheet={name} found {merged_count} merged cell range(s)")

        reported = reported_max_rows[name]
        log(f"sheet={name} reportedMaxRow={reported} — scanning up to 2000 rows for real content...")
        real_last = scan_real_last_row(ws, log=log, sheet_name=name)
        phantom = reported > max(real_last * 3, real_last + 50) and reported > 500

        if phantom:
            any_phantom = True

        reason = []
        if merged_count > 0:
            reason.append(f"{merged_count} merged cells")
        if phantom:
            reason.append(f"reportedMaxRow({reported}) >> realLastRow({real_last})")
        reason_str = ", ".join(reason) if reason else "looks clean"

        log(f"sheet={name} DONE in {_time.time() - t_sheet:.1f}s: realLastRow={real_last}, reportedMaxRow={reported}, merged={merged_count} -> {reason_str}")

        sheets.append({
            "name": name,
            "reportedMaxRow": reported,
            "realLastRow": real_last,
            "mergedCells": merged_count,
            "phantomRows": phantom,
        })

    dirty = total_merged > 0 or any_phantom

    return {
        "dirty": dirty,
        "totalMergedCells": total_merged,
        "sheetCount": len(wb.sheetnames),
        "sheets": sheets,
    }


def clean_workbook(path, out_path, max_data_rows=200):
    import time as _time

    def log(msg):
        print(f"[progress] {msg}", file=sys.stderr, flush=True)

    t_start = _time.time()
    log(f"loading workbook: {path}")
    # IMPORTANT: data_only=True loads formula cells as their last-calculated
    # static value instead of the formula text. Without this, openpyxl
    # preserves the formula string on save but drops Excel's cached result
    # (openpyxl has no calculation engine), leaving AMOUNT/TOTAL AMOUNT
    # columns (which are formulas like =N10*L10) with no readable value at
    # all afterward. Loading as values-only flattens formulas to plain
    # numbers, which is what downstream JSON/dashboard parsing needs anyway.
    wb = openpyxl.load_workbook(path, data_only=True)
    log(f"loaded in {_time.time() - t_start:.1f}s, sheets={wb.sheetnames}")

    per_sheet_summary = []

    for name in wb.sheetnames:
        t_sheet = _time.time()
        ws = wb[name]

        reported_max_row = ws.max_row
        log(f"sheet={name} start reportedMaxRow={reported_max_row}")

        # Identify which columns are PO NUMBER / PR NO. - the dashboard's
        # aggregation logic (buildDashboardData in server.js) uses blank
        # values in these specific columns as its signal for "this is a
        # continuation row within a merged PO group, don't sum it
        # separately." If we fill them in during unmerge like every other
        # column, every continuation row starts looking like its own header
        # row, and the PO's total gets summed multiple times. So: unmerge
        # everything, but skip the value-fill step specifically for these
        # two columns.
        skip_fill_cols = set()
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
            if any(v and "PR DATE" in str(v).upper() for v in row_vals):
                for c, v in enumerate(row_vals, start=1):
                    if v and str(v).strip().upper() in ("PO NUMBER", "PR NO."):
                        skip_fill_cols.add(c)
                break
        if skip_fill_cols:
            log(f"sheet={name} will NOT fill merge values into column(s) {sorted(skip_fill_cols)} (PO NUMBER/PR NO. - dashboard relies on their blankness)")

        # 1. Unmerge every merged range. Copy the anchor cell's VALUE into
        #    every cell in the range (except skip_fill_cols, see above), and
        #    separately copy its STYLE (border/font/fill/alignment) into
        #    every cell in the range unconditionally. Merged "member" cells
        #    normally carry no style of their own - Excel only draws a
        #    border around the merge's outer edge based on the anchor cell,
        #    so once unmerged, the inner cells look borderless/broken unless
        #    we explicitly copy the anchor's style onto them too.
        import copy as _copy
        merged_ranges = list(ws.merged_cells.ranges)
        log(f"sheet={name} unmerging {len(merged_ranges)} ranges...")
        t_unmerge = _time.time()
        for i, merged_range in enumerate(merged_ranges):
            min_col, min_row = merged_range.min_col, merged_range.min_row
            max_col, max_row = merged_range.max_col, merged_range.max_row
            anchor = ws.cell(row=min_row, column=min_col)
            top_left_value = anchor.value
            anchor_border = _copy.copy(anchor.border)
            anchor_font = _copy.copy(anchor.font)
            anchor_fill = _copy.copy(anchor.fill)
            anchor_alignment = _copy.copy(anchor.alignment)
            ws.unmerge_cells(str(merged_range))
            fill_value = min_col not in skip_fill_cols
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    if fill_value:
                        cell.value = top_left_value
                    cell.border = anchor_border
                    cell.font = anchor_font
                    cell.fill = anchor_fill
                    cell.alignment = anchor_alignment
            if (i + 1) % 200 == 0:
                log(f"sheet={name} unmerged {i + 1}/{len(merged_ranges)} ranges ({_time.time() - t_unmerge:.1f}s elapsed)")
        log(f"sheet={name} unmerge done in {_time.time() - t_unmerge:.1f}s")

        # 2. Trim phantom empty rows.
        t_scan = _time.time()
        log(f"sheet={name} scanning up to 2000 rows for real content...")
        real_last = scan_real_last_row(ws, log=log, sheet_name=name)
        log(f"sheet={name} scanned, realLastRow={real_last} ({_time.time() - t_scan:.1f}s)")

        trimmed = 0
        if ws.max_row > real_last:
            trimmed = reported_max_row - real_last
            t_delete = _time.time()
            log(f"sheet={name} deleting {ws.max_row - real_last} phantom rows (this is usually the slow step)...")
            ws.delete_rows(real_last + 1, ws.max_row - real_last)
            log(f"sheet={name} delete_rows done in {_time.time() - t_delete:.1f}s")

        log(f"sheet={name} TOTAL {_time.time() - t_sheet:.1f}s (mergedRemoved={len(merged_ranges)}, phantomTrimmed={trimmed}, finalRows={ws.max_row})")

        per_sheet_summary.append({
            "name": name,
            "mergedCellsRemoved": len(merged_ranges),
            "phantomRowsTrimmed": trimmed,
            "finalRowCount": ws.max_row,
        })

    log(f"saving to {out_path}...")
    t_save = _time.time()
    wb.save(out_path)
    log(f"saved in {_time.time() - t_save:.1f}s. TOTAL RUNTIME {_time.time() - t_start:.1f}s")

    return {
        "ok": True,
        "outPath": out_path,
        "sheets": per_sheet_summary,
    }


def main():
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="mode", required=True)

    p_inspect = sub.add_parser("inspect")
    p_inspect.add_argument("path")

    p_clean = sub.add_parser("clean")
    p_clean.add_argument("path")
    p_clean.add_argument("out_path")

    args = parser.parse_args()

    try:
        if args.mode == "inspect":
            result = inspect_workbook(args.path)
        else:
            result = clean_workbook(args.path, args.out_path)
    except Exception as e:
        print(json.dumps({"ok": False, "error": str(e)}))
        sys.exit(1)

    print(json.dumps(result))
    sys.exit(0)


if __name__ == "__main__":
    main()