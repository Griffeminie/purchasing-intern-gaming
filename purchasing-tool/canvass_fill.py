"""
canvass_fill.py
Fills TEMP.xlsx canvass template with supplied JSON data.
Copies the template, inserts item rows, fills summary fields, saves new file.

Usage: python3 canvass_fill.py <data.json> <template.xlsx> <output.xlsx>
"""
import sys, json, shutil, copy, re, os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.styles.borders import Border, Side

# ── Layout constants ──────────────────────────────────────────────────────────
SUP_COL_WIDTH       = 3    # each supplier block = [label][unit price][total]
FIRST_SUP_COL       = 11   # column K — first supplier block starts here
BUILTIN_SUPS        = 2    # TEMP.xlsx ships with 2 wired-up blocks (K:M, N:P)
NAME_ROW            = 11
INFO_ROW            = 12
FULL_WIDTH_LAST_COL = 16   # P — last column of the template's built-in width


def thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def supplier_cols(idx):
    """(label_col, price_col, total_col) for supplier index idx (0-based).
    label_col is a static caption cell ("PRICE") on item rows and the merged
    anchor cell on header/summary rows — never overwritten with data on item
    rows. price_col holds the real unit price. total_col holds the real total."""
    start = FIRST_SUP_COL + idx * SUP_COL_WIDTH
    return start, start + 1, start + 2


def write_cell(ws, row, col, value):
    """Safely write a value to (row, col). openpyxl only allows writing to
    the TOP-LEFT anchor cell of a merged range — writing to any other cell
    inside that range raises 'MergedCell object attribute value is
    read-only'. If (row, col) turns out to be a non-anchor cell of some
    merge (which can happen after all the row/column shifting this script
    does), redirect the write to that merge's actual anchor instead of
    crashing."""
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= row <= mc.max_row and mc.min_col <= col <= mc.max_col:
            ws.cell(row=mc.min_row, column=mc.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value


def copy_row_style(ws, src_row, dst_row):
    """Copy cell styles from src_row to dst_row."""
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font      = copy.copy(src.font)
            dst.fill      = copy.copy(src.fill)
            dst.border    = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format
    # Also copy row height, so inserted item rows match the template row
    # instead of falling back to Excel's default (this is part of why rows
    # looked randomly too small/tall after export)
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def shift_merges_and_insert_rows(ws, insert_at, n_rows):
    """
    openpyxl's insert_rows() moves cell VALUES down correctly but does NOT
    move MergedCellRange definitions. Any merge at or below `insert_at` has
    to be manually unmerged, the rows inserted, then re-merged at its new
    shifted position — otherwise the old merge boundaries end up overlapping
    the newly written rows, which is what was causing Excel's "repaired /
    removed merge cells" warning and the missing row-15 fields.
    """
    if n_rows <= 0:
        return

    ranges_to_shift = []
    ranges_to_keep  = []
    for mc in list(ws.merged_cells.ranges):
        coords = (mc.min_row, mc.min_col, mc.max_row, mc.max_col)
        if mc.min_row >= insert_at:
            ranges_to_shift.append(coords)
        else:
            ranges_to_keep.append(coords)

    # Unmerge everything first so insert_rows() isn't juggling stale merges
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(start_row=mc.min_row, start_column=mc.min_col,
                          end_row=mc.max_row,   end_column=mc.max_col)

    ws.insert_rows(insert_at, amount=n_rows)

    # Re-apply merges that were above the insertion point — unchanged
    for (r1, c1, r2, c2) in ranges_to_keep:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
# Re-apply merges that were at/below the insertion point — shifted down
    for (r1, c1, r2, c2) in ranges_to_shift:
        ws.merge_cells(start_row=r1 + n_rows, start_column=c1,
                        end_row=r2 + n_rows,   end_column=c2)

def copy_col_block_style(ws, src_start_col, dst_start_col, width, min_row, max_row):
    """Copy cell styles and column widths from a source column block to a
    destination column block, row by row, for the given row range."""
    for r in range(min_row, max_row + 1):
        for offset in range(width):
            src = ws.cell(row=r, column=src_start_col + offset)
            dst = ws.cell(row=r, column=dst_start_col + offset)
            if src.has_style:
                dst.font           = copy.copy(src.font)
                dst.fill           = copy.copy(src.fill)
                dst.border         = copy.copy(src.border)
                dst.alignment      = copy.copy(src.alignment)
                dst.number_format  = src.number_format
    for c in range(width):
        src_letter = get_column_letter(src_start_col + c)
        dst_letter = get_column_letter(dst_start_col + c)
        if src_letter in ws.column_dimensions:
            ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width


def shift_merges_and_insert_cols(ws, insert_at, n_cols):
    """Column equivalent of shift_merges_and_insert_rows — needed when adding
    extra 3-column price blocks for a 3rd+ supplier."""
    if n_cols <= 0:
        return

    ranges_to_shift, ranges_to_keep = [], []
    for mc in list(ws.merged_cells.ranges):
        coords = (mc.min_row, mc.min_col, mc.max_row, mc.max_col)
        (ranges_to_shift if mc.min_col >= insert_at else ranges_to_keep).append(coords)

    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(start_row=mc.min_row, start_column=mc.min_col,
                          end_row=mc.max_row,   end_column=mc.max_col)

    ws.insert_cols(insert_at, amount=n_cols)

    for (r1, c1, r2, c2) in ranges_to_keep:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    for (r1, c1, r2, c2) in ranges_to_shift:
        ws.merge_cells(start_row=r1, start_column=c1 + n_cols,
                        end_row=r2,   end_column=c2 + n_cols)

    # Full-width banners (title bar rows 2-6, "SUPPLIER QUOTATION" header
    # rows 9-10, "NOTHING FOLLOWS" row 15) sit entirely to the left of
    # insert_at so they weren't shifted above — they should stretch to
    # cover the new supplier block. IMPORTANT: this must be restricted to
    # known banner rows, not just "ends at column P" — Supplier 2's own
    # name (row 11), info (row 12), and summary fields (rows 16-23) also
    # happen to end at column P, and widening those by mistake is what was
    # swallowing Supplier 3's columns into Supplier 2's cells.
    BANNER_ROWS = {2, 3, 4, 5, 6, 9, 10, 15}
    for mc in list(ws.merged_cells.ranges):
        if mc.max_col == FULL_WIDTH_LAST_COL and mc.min_col < insert_at and mc.min_row in BANNER_ROWS:
            r1, c1, r2, c2 = mc.min_row, mc.min_col, mc.max_row, mc.max_col
            ws.unmerge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2 + n_cols)


def ensure_supplier_blocks(ws, n_suppliers, bottom_row):
    """Template ships with 2 wired-up 3-column price blocks (K:M, N:P). If more
    suppliers are supplied, clone extra blocks (merges, borders, labels) from
    the 2nd block so the layout scales to any number of suppliers, matching
    the "add as many suppliers as you want" behaviour on the canvass page."""
    extra_needed = max(0, n_suppliers - BUILTIN_SUPS)
    if extra_needed == 0:
        return

    insert_at = FIRST_SUP_COL + BUILTIN_SUPS * SUP_COL_WIDTH  # right after N:P
    n_cols = extra_needed * SUP_COL_WIDTH
    shift_merges_and_insert_cols(ws, insert_at, n_cols)

    src_start = FIRST_SUP_COL + SUP_COL_WIDTH  # N — clone the 2nd block as the template
    for k in range(extra_needed):
        dst_start = insert_at + k * SUP_COL_WIDTH
        copy_col_block_style(ws, src_start, dst_start, SUP_COL_WIDTH, 9, bottom_row)
        # Re-create the per-block merges this new column range needs:
        # name (row 11), info (row 12), and each summary field (rows 16-23).
        for r in (NAME_ROW, INFO_ROW, 16, 17, 18, 19, 20, 21, 22, 23):
            ws.merge_cells(start_row=r, start_column=dst_start,
                            end_row=r,   end_column=dst_start + SUP_COL_WIDTH - 1)

def remove_supplier_block(ws, idx):
    """Completely removes one supplier's 3-column price block (used when
    n_sup == 1, so we don't ship a permanently-empty 'SUPPLIER 2' block).
    Deletes the 3 columns, drops any merge that lived entirely inside them
    (e.g. N21:P21 — supplier 2's own summary cells), shifts merges to the
    right of the block left by 3, and SHRINKS any banner merge that spanned
    across the block (title bar K2:P6, 'SUPPLIER QUOTATION' header G9:P9/
    G10:P10, 'NOTHING FOLLOWS' A20:P20, etc) so it still reaches the new
    last column and stays centered on the new width, instead of leaving a
    dangling 3-column gap or a merge that no longer matches real cells."""
    start, _, _ = supplier_cols(idx)
    delete_at, n_cols = start, SUP_COL_WIDTH
    delete_end = delete_at + n_cols - 1

    def map_col(c):
        if c < delete_at:
            return c
        if c > delete_end:
            return c - n_cols
        return None  # falls inside the removed block

    old_ranges = [(mc.min_row, mc.min_col, mc.max_row, mc.max_col)
                  for mc in ws.merged_cells.ranges]

    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(start_row=mc.min_row, start_column=mc.min_col,
                          end_row=mc.max_row, end_column=mc.max_col)

    # openpyxl doesn't re-key column_dimensions (widths) on delete_cols, so
    # columns to the right of the gap would otherwise keep their OLD width.
    # Capture widths before deleting, then re-apply shifted.
    old_widths = {}
    for c in range(1, 30):
        letter = get_column_letter(c)
        if letter in ws.column_dimensions:
            old_widths[c] = ws.column_dimensions[letter].width

    ws.delete_cols(delete_at, amount=n_cols)

    for c in range(delete_at, 30):
        src_c = c + n_cols
        if src_c in old_widths:
            ws.column_dimensions[get_column_letter(c)].width = old_widths[src_c]

    for (r1, c1, r2, c2) in old_ranges:
        mapped = [m for m in (map_col(c) for c in range(c1, c2 + 1)) if m is not None]
        if not mapped:
            continue  # merge lived entirely inside the removed block — drop it
        new_c1, new_c2 = min(mapped), max(mapped)
        if new_c1 == new_c2 and r1 == r2:
            continue  # collapsed to a single cell — nothing to merge
        ws.merge_cells(start_row=r1, start_column=new_c1, end_row=r2, end_column=new_c2)

def to_number(val):
    """Best-effort parse into a real float. Returns None if it isn't a number
    (blank, or text like 'FREE') — used to keep numeric cells actually numeric
    instead of text, and to avoid #VALUE! errors when doing arithmetic."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if s == "":
        return None
    # Strip currency symbols/labels (₱, P, PHP, Php) and thousands separators
    # before parsing — raw extracted values look like "P 4,550.00" and would
    # otherwise fail float() and silently return None (this was the bug
    # behind unit prices vanishing on some quotations but not others).
    s = re.sub(r"(?i)^(php|₱|p)\s*", "", s)
    s = s.replace(",", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None
    
def sanitize_sheet_name(name):
    """Excel sheet names: max 31 chars, can't contain : \\ / ? * [ ]."""
    for ch in r'[]:\/?*':
        name = name.replace(ch, '-')
    name = name.strip() or 'Canvass'
    return name[:31]


def unique_sheet_name(wb, name):
    """Avoids clobbering an existing tab with the same name — e.g. two
    HR canvasses both titled 'Office Supplies' become 'Office Supplies'
    and 'Office Supplies (2)'."""
    name = sanitize_sheet_name(name)
    if name not in wb.sheetnames:
        return name
    base = name[:27]  # leave room for " (n)"
    n = 2
    while True:
        candidate = f"{base} ({n})"[:31]
        if candidate not in wb.sheetnames:
            return candidate
        n += 1


def copy_sheet_into_workbook(src_ws, dst_wb, sheet_name):
    """Manually clones src_ws (values, styles, merges, row/col sizing) into
    a brand-new sheet inside dst_wb. openpyxl worksheets belong to exactly
    one workbook, so there's no built-in 'move this sheet to another file' —
    everything has to be copied cell by cell instead of the object moved."""
    name = unique_sheet_name(dst_wb, sheet_name)
    dst_ws = dst_wb.create_sheet(title=name)  # created at the end by default

    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font          = copy.copy(cell.font)
                new_cell.fill          = copy.copy(cell.fill)
                new_cell.border        = copy.copy(cell.border)
                new_cell.alignment     = copy.copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection    = copy.copy(cell.protection)

    for mc in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(start_row=mc.min_row, start_column=mc.min_col,
                            end_row=mc.max_row,   end_column=mc.max_col)

    for letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[letter].width = dim.width
    for idx, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst_ws.row_dimensions[idx].height = dim.height

    # create_sheet() already appends at the end, but move it explicitly so
    # this stays correct even if a same-named "(2)" tab gets re-added later.
    dst_wb.move_sheet(name, offset=len(dst_wb.sheetnames) - 1 - dst_wb.sheetnames.index(name))
    return dst_ws


def main():
    data_path, tmpl_path, out_path = sys.argv[1], sys.argv[2], sys.argv[3]
    # Optional: append this canvass as a new LAST sheet into an existing
    # multi-department workbook instead of shipping it as its own file.
    #   argv[4] = path to the existing book (e.g. HR_Canvass_Book.xlsx)
    #   argv[5] = desired sheet/tab name (defaults to the canvass title)
    append_target = sys.argv[4] if len(sys.argv) > 4 else None
    sheet_name    = sys.argv[5] if len(sys.argv) > 5 else None

    with open(data_path) as f:
        d = json.load(f)

    # ── Unpack payload ────────────────────────────────────────────────────────
    title      = d.get("title", "")
    ro         = d.get("ro", "")
    date_val   = d.get("date", "")
    remarks    = d.get("remarks", "")
    items      = d.get("items", [])   # [{desc, qty, unit, p1, p2, p3}, ...]

    suppliers  = d.get("suppliers", [])   # [{name, loc, contact, num}, ...] — any length
    n_sup      = max(len(suppliers), 1)

    dl         = d.get("delivery", [])
    discount   = d.get("discount", [])
    credit     = d.get("credit",   [])
    vat        = d.get("vat",      [])
    avail      = d.get("avail",    [])
    da         = d.get("da",       [])

    def sup_val(lst, i, default=""):
        v = lst[i] if i < len(lst) else None
        return v if v not in (None, "") else default

    # ── Copy template ─────────────────────────────────────────────────────────
    shutil.copy2(tmpl_path, out_path)
    wb = load_workbook(out_path)
    ws = wb.active  # Sheet1

    # If there's only ONE supplier, drop the template's unused 2nd price
    # block (columns N:P) entirely instead of leaving it blank — shrinks and
    # re-centers every banner merge that spanned across it.
    if n_sup == 1:
        remove_supplier_block(ws, 1)

    # Sheet/tab name = the project title, falling back to "Canvass" when
    # blank. Applies to standalone exports; append-mode uses the separate
    # sheet_name/title fallback further down when cloning into the master book.
    ws.title = sanitize_sheet_name(title) if title else "Canvass"

    # ── Fill header fields ────────────────────────────────────────────────────
    # Row 2: item/title
    write_cell(ws, 2, 11, title or "-ITEM NAME OR SUPPLIER TITLE HERE-")

    # Row 13: title + RO (A13:F13 merged)
    title_ro = title
    if ro: title_ro += f"\nRO: {ro}"
    else:  title_ro += "\nRO: IF NO RO LEAVE BLANK"
    write_cell(ws, 13, 1, title_ro)

    # ── Make sure enough supplier price-blocks physically exist ──────────────
    # (clones extra K:M-style blocks for supplier 3, 4, 5... if needed)
    ensure_supplier_blocks(ws, n_sup, bottom_row=23)

    for i, sup in enumerate(suppliers):
        label_col, _, _ = supplier_cols(i)
        # NAME_ROW (row 11) is the template's own static "SUPPLIER 1" /
        # "SUPPLIER 2" caption — never overwrite it with the actual company
        # name. The real company name goes as the FIRST line of the info
        # cell below it (row 12), together with location/contact/number,
        # so "SUPPLIER 1" always stays visible as the block's fixed label.
        info = "\n".join(filter(None, [
            sup.get("name", ""),
            sup.get("loc", ""),
            sup.get("contact", ""),
            sup.get("num", ""),
        ])) or "Company Name\nLocation\nContact person\nTheir Number"
        write_cell(ws, INFO_ROW, label_col, info)

    # ── Insert item rows above row 15 (NOTHING FOLLOWS) ──────────────────────
    # Template has 1 sample row at row 14. We'll overwrite it with real data
    # and insert extras as needed.

    n_items = len(items)
    if n_items == 0:
        items = [{"desc": "", "qty": 1, "unit": "pce", "p1": "", "p2": "", "p3": ""}]
        n_items = 1

    # Insert (n_items - 1) extra rows after row 14 by shifting rows down
    if n_items > 1:
        shift_merges_and_insert_rows(ws, 15, n_items - 1)
        # Copy style from row 14 to the new rows
        for extra in range(1, n_items):
            copy_row_style(ws, 14, 14 + extra)
            # B:D — item description. This merge is new (row 14 is the
            # only row that had it originally) so it's created fresh
            # rather than shifted.
            ws.merge_cells(
                start_row=14 + extra, start_column=2,
                end_row=14 + extra,   end_column=4
            )

        # Column J (row 11 through the original sample row 14) is ONE
        # continuous decorative vertical bar in the template — J11:J14 —
        # NOT a per-row I:J merge. shift_merges_and_insert_rows() leaves it
        # untouched at rows 11-14 (its min_row is above the insertion
        # point), so it needs to be manually stretched down to keep
        # covering every newly inserted item row, instead of stopping
        # short at row 14 while new rows sit unmerged/uncovered below it.
        for mc in list(ws.merged_cells.ranges):
            if (mc.min_col == 10 and mc.max_col == 10
                    and mc.min_row == NAME_ROW and mc.max_row == 14):
                ws.unmerge_cells(start_row=mc.min_row, start_column=mc.min_col,
                                  end_row=mc.max_row, end_column=mc.max_col)
                ws.merge_cells(start_row=NAME_ROW, start_column=10,
                                end_row=14 + n_items - 1, end_column=10)
                break

    # ── Fill item rows ────────────────────────────────────────────────────────
    sup_totals = [0.0] * n_sup
    last_col   = FIRST_SUP_COL + n_sup * SUP_COL_WIDTH - 1

    for idx, item in enumerate(items):
        r = 14 + idx
        desc    = item.get("desc", "")
        qty     = item.get("qty", 1)
        unit    = item.get("unit", "pce")
        qty_num = to_number(qty)

        ws.cell(row=r, column=1).value  = idx + 1                                  # Item No
        ws.cell(row=r, column=2).value  = desc                                     # Description (B, merged to D)
        ws.cell(row=r, column=5).value  = qty_num if qty_num is not None else qty  # QTY as real number
        ws.cell(row=r, column=6).value  = unit                                     # UNIT

        # Reference Only (Last Purchase) — columns H (Unit Price) and I
        # (Date), one per item row, sitting to the left of every supplier
        # price block. Column J is the decorative vertical bar next to it
        # and is never written to directly.
        last_price = item.get("lastPrice", "")
        last_price_num = to_number(last_price)
        ws.cell(row=r, column=8).value = last_price_num if last_price_num is not None else last_price
        ws.cell(row=r, column=9).value = item.get("lastDate", "")

        for i in range(n_sup):
            label_col, price_col, total_col = supplier_cols(i)
            raw_price = item.get(f"p{i + 1}", "")
            price_num = to_number(raw_price)

            if price_num is None and raw_price not in ("", None):
                print(f"[canvass_fill] WARNING row {r} supplier {i+1}: "
                      f"could not parse price {raw_price!r} -> writing blank cell "
                      f"(col {get_column_letter(price_col)})", file=sys.stderr)

            # label_col (K, N, Q, ...) is the static "PRICE" caption baked
            # into the template on item rows — never written to. Only
            # price_col gets the real unit price, only total_col gets the
            # real total formula.
            ws.cell(row=r, column=price_col).value = price_num
            if price_num is not None:
                price_letter = get_column_letter(price_col)
                ws.cell(row=r, column=total_col).value = f"={price_letter}{r}*E{r}"
                sup_totals[i] += price_num * (qty_num or 0)
            else:
                print(f"[canvass_fill] row {r} supplier {i+1}: total cell "
                      f"NOT written (price_num is None)", file=sys.stderr)

        # Apply borders
        for col in range(1, last_col + 1):
            ws.cell(row=r, column=col).border = thin()

    # After inserting rows, NOTHING FOLLOWS is at row 14 + n_items
    nf_row = 14 + n_items
    # Summary rows shift accordingly — fill them
    # Row offsets from NOTHING FOLLOWS (nf_row):
    #   nf_row    = NOTHING FOLLOWS
    #   nf_row+1  = Remarks / DELIVERY CHARGE
    #   nf_row+2  = TOTAL
    #   nf_row+3  = DISCOUNT
    #   nf_row+4  = GRAND TOTAL
    #   nf_row+5  = CREDIT TERMS
    #   nf_row+6  = VAT
    #   nf_row+7  = AVAILABILITY
    #   nf_row+8  = DELIVERY ADDRESS

    def sfill(row_offset, col, value):
        write_cell(ws, nf_row + row_offset, col, value)

    for i in range(n_sup):
        label_col, _, _ = supplier_cols(i)  # summary rows are merged K:M-style — write to the anchor (label_col)

        delivery_i = to_number(sup_val(dl, i))
        discount_i = to_number(sup_val(discount, i))
        total_i    = sup_totals[i]
        grand_i    = total_i + (delivery_i or 0) - (discount_i or 0)

        # Delivery charge — real number if one was given, otherwise "FREE"
        sfill(1, label_col, delivery_i if delivery_i is not None else (sup_val(dl, i) or "FREE"))
        # Total — plain computed number (calculated in Python, not a live SUM formula)
        sfill(2, label_col, total_i)
        # Discount
        sfill(3, label_col, discount_i if discount_i is not None else sup_val(discount, i))
        # Grand total — plain computed number, never a formula, so it can
        # never throw #VALUE! no matter what text ends up in delivery/discount
        sfill(4, label_col, grand_i)
        # Credit terms
        sfill(5, label_col, sup_val(credit, i))
        # VAT
        sfill(6, label_col, sup_val(vat, i, "VAT INCLUSIVE"))
        # Availability
        sfill(7, label_col, sup_val(avail, i))
        # Delivery address
        sfill(8, label_col, sup_val(da, i))

    # Remarks (A16:C23 merged area)
    sfill(1, 1, remarks or "")

    # Defensive safeguard: re-assert every image's anchor exactly as loaded
    # from the template, immediately before saving. The logo lives at
    # A1:C6, entirely outside the supplier price blocks we insert/delete/
    # re-merge columns and rows around — it should never move — but this
    # guards against any upstream operation subtly disturbing an anchor.
    _preserved_images = list(ws._images)
    ws._images = _preserved_images

    if append_target:
        if os.path.exists(append_target):
            master_wb = load_workbook(append_target)
        else:
            # Target book doesn't exist yet (e.g. first-ever HR canvass) —
            # start a fresh one without openpyxl's default blank "Sheet".
            master_wb = Workbook()
            master_wb.remove(master_wb.active)

        final_name = sheet_name or title or "Canvass"
        copy_sheet_into_workbook(ws, master_wb, final_name)
        # IMPORTANT: save back to append_target itself, in place — NOT to
        # out_path, which is just the scratch file used to build this one
        # sheet. Saving to out_path here would silently discard the append
        # and leave the real department book untouched.
        master_wb.save(append_target)
        print(f"Appended sheet '{final_name}' to {append_target} (saved in place)")
    else:
        wb.save(out_path)
        print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()