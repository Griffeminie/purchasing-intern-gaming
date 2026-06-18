"""
export_canvass.py
-----------------
Fills TEMP.xlsx (the canvass template) with data from the web app and
saves a new file.  Supports 1-3 suppliers (template has columns for 2;
a 3rd supplier block is appended when needed).

Usage (called by your backend API endpoint):
    python export_canvass.py '<json_payload>' output.xlsx

JSON payload shape (matches canvass.html exportToTemplate()):
{
  "title":    "Office Supplies Q3",
  "ro":       "RO-001",
  "date":     "2026-06-18",
  "remarks":  "...",
  "items": [
    { "desc":"Ballpen", "qty":10, "unit":"pce",
      "prices": [150, 140, 155] }   // one price per supplier, in order
  ],
  "suppliers": [
    { "name":"ABC Corp", "loc":"Makati", "contact":"Juan", "num":"09xx",
      "dl":"FREE", "disc":"", "ct":"30 days", "vat":"VAT Inclusive",
      "av":"3 days", "da":"123 Main St" },
    ...
  ]
}
"""

import sys, json, copy, shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE = "TEMP.xlsx"

# ── Column layout for 2-supplier template ────────────────────────────────────
# Supplier 1:  unit-price=K(11)  total=M(13)
# Supplier 2:  unit-price=N(14)  total=P(16)
# Each supplier block = 3 columns: unit-price, unit-price-display, total
SUP_COLS = [
    {"up": 11, "upd": 12, "tot": 13},   # S1  K L M
    {"up": 14, "upd": 15, "tot": 16},   # S2  N O P
    {"up": 17, "upd": 18, "tot": 19},   # S3  Q R S  (added if needed)
]

thin = Side(style="thin")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill_template(payload: dict, output_path: str):
    shutil.copy(TEMPLATE, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    suppliers = payload.get("suppliers", [])
    items     = payload.get("items", [])
    n_sup     = len(suppliers)

    # ── 1. Header cells ───────────────────────────────────────────────────────
    title_str = payload.get("title", "")
    ro_str    = payload.get("ro", "")
    title_line = title_str + ("\nRO: " + ro_str if ro_str else "")
    ws["A13"] = title_line or "TITLE HERE"

    date_str = payload.get("date", "")
    ws["K2"] = title_str or "-ITEM NAME OR SUPPLIER TITLE HERE-"

    # Date goes in the I13 area (the " Date" header column)
    ws["I13"] = date_str

    # ── 2. Supplier header info (rows 11-12) ──────────────────────────────────
    for i, sup in enumerate(suppliers[:3]):
        col_cfg = SUP_COLS[i]
        col_ltr = get_column_letter(col_cfg["up"])
        # The merged cell K12 / N12 / Q12
        cell = ws[f"{col_ltr}12"]
        cell.value = "\n".join(filter(None, [
            sup.get("name",""), sup.get("loc",""),
            sup.get("contact",""), sup.get("num","")
        ]))

        # Supplier label row 11  (K11, N11, Q11)
        if i < 2:
            pass  # already in template
        else:
            # 3rd supplier: write label manually
            ws[f"{col_ltr}11"] = f"SUPPLIER {i+1}"

    # ── 3. Item rows (starting at row 14) ────────────────────────────────────
    START_ROW   = 14
    NOTHING_ROW = START_ROW + len(items)  # "NOTHING FOLLOWS" row

    # Copy row-14 template cell styles for new rows
    template_row_cells = {col: ws.cell(START_ROW, col) for col in range(1, 20)}

    def copy_style(src_cell, dst_cell):
        if src_cell.font:
            dst_cell.font = copy.copy(src_cell.font)
        if src_cell.border:
            dst_cell.border = copy.copy(src_cell.border)
        if src_cell.fill and src_cell.fill.patternType:
            dst_cell.fill = copy.copy(src_cell.fill)
        if src_cell.alignment:
            dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

    for idx, item in enumerate(items):
        row = START_ROW + idx
        if idx > 0:
            ws.insert_rows(row)
            ws.row_dimensions[row].height = ws.row_dimensions[START_ROW].height or 36

        prices = item.get("prices", [])
        # Fill merged cells by writing to first cell of merge
        ws.cell(row, 1).value = idx + 1               # Item no.
        ws.cell(row, 2).value = item.get("desc", "")  # Description (B:D merged)
        ws.cell(row, 5).value = item.get("qty", 1)    # Qty
        ws.cell(row, 6).value = item.get("unit", "pce")  # Unit

        for i in range(min(n_sup, 3)):
            price = prices[i] if i < len(prices) else ""
            cfg   = SUP_COLS[i]
            up_col  = cfg["up"]
            upd_col = cfg["upd"]
            tot_col = cfg["tot"]
            qty_col = 5

            ws.cell(row, up_col).value  = price if price != "" else ""
            ws.cell(row, upd_col).value = price if price != "" else ""
            if price != "":
                ws.cell(row, tot_col).value = \
                    f"={get_column_letter(up_col)}{row}*{get_column_letter(qty_col)}{row}"
            else:
                ws.cell(row, tot_col).value = ""

        # Apply styles from template row
        if idx > 0:
            for col in range(1, 20):
                copy_style(template_row_cells[col], ws.cell(row, col))
                ws.cell(row, col).alignment = Alignment(
                    horizontal=ws.cell(START_ROW, col).alignment.horizontal or "center",
                    vertical="center", wrap_text=True
                )

    # ── 4. "NOTHING FOLLOWS" row ──────────────────────────────────────────────
    nf_row = START_ROW + len(items)
    ws.cell(nf_row, 1).value = "**NOTHING FOLLOWS**"

    # ── 5. Remarks + terms rows (shift if rows were inserted) ─────────────────
    # Because we inserted rows, we need to recalculate where the footer is.
    # The original footer starts at what was row 15 (NOTHING FOLLOWS) +1 = 16.
    # After inserts, footer base = nf_row + 1
    footer_base = nf_row + 1

    remarks_row  = footer_base      # was 16
    total_row    = footer_base + 1  # was 17
    disc_row     = footer_base + 2  # was 18
    gtotal_row   = footer_base + 3  # was 19
    ct_row       = footer_base + 4  # was 20
    vat_row      = footer_base + 5  # was 21
    av_row       = footer_base + 6  # was 22
    da_row       = footer_base + 7  # was 23

    def set_footer(row, s_idx, value):
        col_ltr = get_column_letter(SUP_COLS[s_idx]["up"])
        ws[f"{col_ltr}{row}"] = value

    ws[f"A{remarks_row}"] = "Remarks: " + payload.get("remarks", "")

    for i, sup in enumerate(suppliers[:3]):
        # Total formula
        if len(items) > 0:
            tot_col = get_column_letter(SUP_COLS[i]["tot"])
            ws.cell(total_row, SUP_COLS[i]["up"]).value = \
                f"=SUM({tot_col}{START_ROW}:{tot_col}{nf_row-1})"
        set_footer(disc_row,   i, sup.get("disc", ""))
        # Grand total = total + delivery - discount (simple label; formula if both numeric)
        dl_val   = sup.get("dl", "FREE")
        disc_val = sup.get("disc", "")
        ws.cell(gtotal_row, SUP_COLS[i]["up"]).value = \
            f"=IF({get_column_letter(SUP_COLS[i]['up'])}{total_row}=\"\",\"\"," \
            f"{get_column_letter(SUP_COLS[i]['up'])}{total_row})"
        set_footer(ct_row,   i, sup.get("ct", ""))
        set_footer(vat_row,  i, sup.get("vat", ""))
        set_footer(av_row,   i, sup.get("av", ""))
        set_footer(da_row,   i, sup.get("da", ""))
        # Delivery charge
        ws.cell(remarks_row + 0, SUP_COLS[i]["up"]).value  # keep remarks
        ws.cell(footer_base, SUP_COLS[i]["up"]).value       # delivery at remarks_row
        # Delivery is on the same row as remarks
        ws.cell(remarks_row, SUP_COLS[i]["up"]).value = dl_val

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python export_canvass.py '<json>' output.xlsx")
        sys.exit(1)
    payload = json.loads(sys.argv[1])
    fill_template(payload, sys.argv[2])
